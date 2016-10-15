# python 3.5.2

## DESCRIPTION: this project implements the "Multi-class classification-based algorithm"
# to create a software which identifies the neuron cell type based on specific input cell characteristics. 




#package name (lib)
import openpyxl # openpyxl allows python to open/read excel files.

# fnction to read excel file and return a list of values in cell renage
# input: file_name is a string representing name of xls file to read e.g. 'ram.xlsx'
#        cell_range is a string representing range of cells to be read e.g. "A5:A10"
# output: data a list containing the values of the cells specified e.g. values of A5:A10

def read_file(file_name,cell_range):
    # check whether file_name,cell_range is empty or not
    if len (file_name) == 0:
        return None
    elif len(cell_range) == 0:
        return None  
    else:
        # wb is a variable which uses openpyxl library
        # to run the function load_workbook(file_name)
        wb = openpyxl.load_workbook(file_name)
        # sheet is a variable which brings the active sheet from wb.    
        sheet=wb.active
        # declaring an empty list to hold the cells
        data = []

        # extract cells from sheet
        if cell_range is "all":
            for index, row in enumerate(sheet.iter_rows()):
                data.append([])
                for cell in row:
                    data[index].append(cell.value)
        else:
            for cell in sheet[cell_range]:  
                data.append(cell[0].value)
        # return the cells 
        return data

# function to transform rows to columns.
# it returns a dictionary in which keys are the col names, and values are list of col values
# myrowdata is an argument which should be a list of list
# if the argument is not a list of list then a None value is returned 
# the 1st list of the argument is treated as col headers for the return dictionary 

def transform_row_to_col(myrowdata):
    # M16.tc1: the test case for this is test_passing_string_as_row_data. exp output = None 
    if type(myrowdata) is list:
        # first row has to be col header (this is a must requirment)
        all_col_header = myrowdata[0] 

        # M16.tc2:test_passing_just_a_list_as_row_data. exp output = None
        if type(all_col_header) is list: 

            # M16.tc3: Do the transformation from row to col
            # creating an empty dictionary to store transformed rows into cols
            mycoldata = {}
            # The ourter loop is navigating thru columns of the list i.e. moving in the x axis of the sheet
            for col_index in range(0, len(all_col_header)):
                # accessing col header based on col_index
                col_hdr = all_col_header[col_index] 

                # putting an empty list into the dictionary to store the col values
                mycoldata[col_hdr] = []

                # the inner loop is navigating thru rows i.e. moving in the y axsis of the sheet 
                for row_index in range(1, len(myrowdata)):
                    #accessing value based on row_index
                    value_list = myrowdata[row_index]
                    # appending the cell values to the col list i.e. adding col values row by row 
                    mycoldata[col_hdr].append(value_list[col_index])
            # returning the ouput
            return mycoldata

        else: #this is checking the sublist  if type(all_col_header) is list
            return None
        
    else: #this is checking the outer list if type(myrowdata) is list
        return None

# creating a function to extract data set from row data
# depending on row_filter_list and col_filter_list
# refer to test_read_data.py (M23,M24) for test case of this function.
# arguments for this function must be in form of list. 
def extract_ds_from_row_data(myrowdata,row_filter_list,col_filter_list):
        
    # if type(myrowdata,row_filter_list,col_filter_list) are lists, then return myetd
    if type(myrowdata) is list:       
        if type(col_filter_list) is list:
        ### if type(row_filter_list) is list:   ---> this if statement returns an error. why?                 
            # myetd is a variable to hold extracted data
            myetd = []
            # creating an outer loop to navigate rows as per the filterd list
            for row_index in row_filter_list:
                # myerow is a variable to hold extracted rows 
                myerow = []
                # mycurrentrow is a variable to hold current row from given data set 
                mycurrentrow = myrowdata[row_index]
                # creating an innter loop to navigate column as per filtered list
                for col_index in col_filter_list:
                    # extracting values from current row column wise
                    myerow.append(mycurrentrow[col_index])

                # appending extracted row into output dataset
                myetd.append(myerow)
            # return output
            return myetd
    else:
        return None
